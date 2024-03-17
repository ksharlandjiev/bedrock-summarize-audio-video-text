from openpyxl import load_workbook
from handlers.abstract_handler import AbstractHandler

class MicrosoftExcelReaderHandler(AbstractHandler):

    def handle(self, request: dict) -> dict:
        print("Processing Excel file...")

        # Initialize a variable to hold the extracted text
        text_content = ''

        # Load the workbook and select the first worksheet
        workbook = load_workbook(filename=request.get("path", None))
        worksheet = workbook.active

        # Iterate through each row and cell to extract text
        for row in worksheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:  # Check if the cell is not empty
                    text_content += str(cell) + ' '
            text_content += '\n'  # Add a newline after each row

        # Update the request with the extracted text
        request.update({"text": text_content})

        # Call the next handler in the chain
        return super().handle(request)
